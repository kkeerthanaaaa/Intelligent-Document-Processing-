"""
DocIntel — Results Router v2.1
Bug fix: images_json parse operator precedence fixed.
All fields parsed with explicit defaults.
"""

import json, csv, io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import aiosqlite
from backend.database import DB_PATH

router = APIRouter()

JSON_LIST_FIELDS = ("tables_json", "entities_json", "images_json")
JSON_DICT_FIELDS = ("kv_json",)
JSON_OBJ_FIELDS  = ("metadata_json",)


def _parse_result_row(d: dict) -> dict:
    """
    Correctly parse all JSON fields with proper defaults.
    Fix: operator precedence bug — must use explicit if/else, not inline ternary.
    """
    for field in JSON_LIST_FIELDS:
        raw = d.get(field) or "[]"
        try:
            parsed = json.loads(raw)
            d[field] = parsed if isinstance(parsed, list) else []
        except Exception:
            d[field] = []

    for field in JSON_DICT_FIELDS:
        raw = d.get(field) or "{}"
        try:
            parsed = json.loads(raw)
            d[field] = parsed if isinstance(parsed, dict) else {}
        except Exception:
            d[field] = {}

    for field in JSON_OBJ_FIELDS:
        raw = d.get(field) or "{}"
        try:
            d[field] = json.loads(raw)
        except Exception:
            d[field] = {}

    return d


@router.get("")
async def list_results():
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id, job_id, filename, doc_type, confidence, page_count, "
            "summary, images_json, created_at "
            "FROM results ORDER BY created_at DESC LIMIT 200"
        ) as cur:
            rows = await cur.fetchall()

    results = []
    for row in rows:
        d = dict(row)
        # Parse images_json to get count
        try:
            imgs = json.loads(d.get("images_json") or "[]")
            d["images_count"] = len(imgs) if isinstance(imgs, list) else 0
        except Exception:
            d["images_count"] = 0
        # Don't send full images_json in list — too large
        d.pop("images_json", None)
        results.append(d)

    return {"results": results}


@router.get("/{job_id}")
async def get_result(job_id: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM results WHERE job_id=?", (job_id,)
        ) as cur:
            row = await cur.fetchone()

    if not row:
        raise HTTPException(404, "Result not found — job may still be processing")

    d = _parse_result_row(dict(row))
    return d


@router.get("/{job_id}/export/json")
async def export_json(job_id: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM results WHERE job_id=?", (job_id,)
        ) as cur:
            row = await cur.fetchone()

    if not row:
        raise HTTPException(404, "Result not found")

    d     = _parse_result_row(dict(row))
    fname = (d.get("filename") or job_id).rsplit(".", 1)[0] + "_result.json"
    return StreamingResponse(
        io.StringIO(json.dumps(d, indent=2, ensure_ascii=False)),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get("/{job_id}/export/csv")
async def export_csv(job_id: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM results WHERE job_id=?", (job_id,)
        ) as cur:
            row = await cur.fetchone()

    if not row:
        raise HTTPException(404, "Result not found")

    d      = _parse_result_row(dict(row))
    kv     = d.get("kv_json", {})
    tables = d.get("tables_json", [])
    buf    = io.StringIO()
    w      = csv.writer(buf)

    w.writerow(["Field", "Value"])
    w.writerow(["Filename",      d.get("filename", "")])
    w.writerow(["Document Type", d.get("doc_type", "")])
    w.writerow(["Confidence",    f"{d.get('confidence', '')}%"])
    w.writerow(["Pages",         d.get("page_count", "")])
    w.writerow(["Processed At",  d.get("created_at", "")])
    w.writerow([])
    w.writerow(["=== EXTRACTED KEY-VALUES ===", ""])
    for k, v in kv.items():
        w.writerow([k, v])
    w.writerow([])
    for ti, tbl in enumerate(tables):
        w.writerow([f"=== TABLE {ti+1}: {tbl.get('title', '')} ===", ""])
        w.writerow(tbl.get("headers", []))
        for row2 in tbl.get("rows", []):
            w.writerow(row2)
        w.writerow([])
    w.writerow(["=== SUMMARY ===", ""])
    w.writerow(["", d.get("summary", "")])

    buf.seek(0)
    fname = (d.get("filename") or job_id).rsplit(".", 1)[0] + "_data.csv"
    return StreamingResponse(
        buf, media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get("/{job_id}/export/excel")
async def export_excel(job_id: str):
    """
    Export full result as a formatted Excel workbook.
    Sheet 1: Document Info + Key-Values
    Sheet 2+: Each extracted table on its own sheet (with full row/column data)
    Sheet N: All Entities
    Last: Full Text
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM results WHERE job_id=?", (job_id,)
        ) as cur:
            row = await cur.fetchone()

    if not row:
        raise HTTPException(404, "Result not found")

    d        = _parse_result_row(dict(row))
    kv       = d.get("kv_json", {})
    tables   = d.get("tables_json", [])
    entities = d.get("entities_json", [])
    images   = d.get("images_json", [])
    summary  = d.get("summary", "")
    filename = d.get("filename", job_id)

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────
    BLUE_FILL  = PatternFill("solid", fgColor="1D4ED8")
    GREEN_FILL = PatternFill("solid", fgColor="059669")
    ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")
    ALT2_FILL  = PatternFill("solid", fgColor="F0FDF4")
    WH_FONT    = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1D4ED8")
    LABEL_FONT = Font(bold=True, size=10)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP       = Alignment(wrap_text=True, vertical="top")
    THIN       = Side(style="thin", color="D4D0C8")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr_row(ws, row_num, num_cols, fill=None):
        f = fill or BLUE_FILL
        for c in range(1, num_cols + 1):
            cell = ws.cell(row_num, c)
            cell.fill = f
            cell.font = WH_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def data_row(ws, row_num, num_cols, alt=False, fill=None):
        f = fill if fill else (ALT_FILL if alt else None)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row_num, c)
            if f: cell.fill = f
            cell.border = BORDER
            cell.alignment = WRAP

    def auto_col(ws, min_w=10, max_w=55):
        for col in ws.columns:
            mx = 0
            cl = get_column_letter(col[0].column)
            for cell in col:
                try: mx = max(mx, min(len(str(cell.value or "")), max_w))
                except: pass
            ws.column_dimensions[cl].width = max(min_w, min(mx + 4, max_w))

    # ── Sheet 1: Document Info & Key-Values ─────────────
    ws1 = wb.active
    ws1.title = "Document Info"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"DocIntel Extraction Report — {filename}"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:D2")
    ws1["A2"] = (f"Type: {d.get('doc_type','')}  |  "
                 f"Confidence: {d.get('confidence','')}%  |  "
                 f"Pages: {d.get('page_count','')}  |  "
                 f"Images: {len(images)}  |  "
                 f"Tables: {len(tables)}  |  "
                 f"Processed: {d.get('created_at','')}")
    ws1["A2"].font = Font(size=9, color="6B6860")
    ws1["A2"].alignment = CENTER
    ws1.row_dimensions[2].height = 18

    # KV headers
    for col, txt in enumerate(["Field", "Value", "Field", "Value"], 1):
        ws1.cell(4, col, txt)
    hdr_row(ws1, 4, 4)
    ws1.row_dimensions[4].height = 22

    kv_items = list(kv.items())
    half = (len(kv_items) + 1) // 2
    for i, (k, v) in enumerate(kv_items[:half]):
        r = 5 + i
        c = ws1.cell(r, 1, k); c.font = LABEL_FONT; c.border = BORDER
        c2 = ws1.cell(r, 2, v); c2.border = BORDER
        if i % 2 == 1:
            c.fill = ALT_FILL; c2.fill = ALT_FILL
        ws1.row_dimensions[r].height = 18
    for i, (k, v) in enumerate(kv_items[half:]):
        r = 5 + i
        c = ws1.cell(r, 3, k); c.font = LABEL_FONT; c.border = BORDER
        c2 = ws1.cell(r, 4, v); c2.border = BORDER
        if i % 2 == 1:
            c.fill = ALT_FILL; c2.fill = ALT_FILL

    # Summary
    sr = 5 + max(half, len(kv_items) - half) + 2
    ws1.merge_cells(f"A{sr}:D{sr}")
    ws1[f"A{sr}"] = "SUMMARY"
    ws1[f"A{sr}"].font = WH_FONT
    ws1[f"A{sr}"].fill = BLUE_FILL
    ws1[f"A{sr}"].alignment = CENTER
    ws1.row_dimensions[sr].height = 22
    sr += 1
    ws1.merge_cells(f"A{sr}:D{sr+2}")
    ws1[f"A{sr}"] = summary
    ws1[f"A{sr}"].alignment = WRAP
    ws1[f"A{sr}"].border = BORDER
    ws1.row_dimensions[sr].height = 60

    auto_col(ws1)

    # ── Sheets 2+: Each table ────────────────────────────
    for ti, tbl in enumerate(tables):
        title   = tbl.get("title", f"Table {ti+1}")
        headers = tbl.get("headers", [])
        rows    = tbl.get("rows", [])
        if not rows:
            continue
        ncols   = max(len(headers), max((len(r) for r in rows), default=1))
        sheet_nm = title[:28].replace("/","-").replace("\\","-")

        ws = wb.create_sheet(title=sheet_nm)
        ws.sheet_view.showGridLines = False

        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        ws["A2"] = (f"Source: {filename}  |  "
                    f"Page: {tbl.get('page','?')}  |  "
                    f"{len(rows)} row(s)  |  "
                    f"{ncols} column(s)")
        ws["A2"].font = Font(size=9, color="6B6860")
        ws["A2"].alignment = CENTER
        ws.row_dimensions[2].height = 16

        if headers:
            for ci, h in enumerate(headers, 1):
                ws.cell(3, ci, h)
            hdr_row(ws, 3, ncols, fill=GREEN_FILL)
            ws.row_dimensions[3].height = 24

        start = 4
        for ri, row_data in enumerate(rows):
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(start + ri, ci, val)
                c.border = BORDER
                c.alignment = WRAP
                if ri % 2 == 1:
                    c.fill = ALT2_FILL
            ws.row_dimensions[start + ri].height = 20

        ws.freeze_panes = ws["A4"]
        auto_col(ws)

    # ── Entities sheet ───────────────────────────────────
    if entities:
        wse = wb.create_sheet("Entities")
        wse.sheet_view.showGridLines = False
        wse.merge_cells("A1:C1")
        wse["A1"] = "Named Entities"
        wse["A1"].font = TITLE_FONT
        wse["A1"].alignment = CENTER
        wse.row_dimensions[1].height = 28
        for ci, h in enumerate(["Entity Text", "Type", "Position"], 1):
            wse.cell(2, ci, h)
        hdr_row(wse, 2, 3)
        for ri, ent in enumerate(entities, 3):
            wse.cell(ri, 1, ent.get("text", ""))
            wse.cell(ri, 2, ent.get("type", ""))
            wse.cell(ri, 3, f"{ent.get('start','')}–{ent.get('end','')}")
            data_row(wse, ri, 3, alt=(ri % 2 == 0))
            wse.row_dimensions[ri].height = 16
        wse.freeze_panes = wse["A3"]
        auto_col(wse)

    # ── Images sheet ─────────────────────────────────────
    if images:
        wsi = wb.create_sheet("Images")
        wsi.sheet_view.showGridLines = False
        wsi.merge_cells("A1:E1")
        wsi["A1"] = "Extracted Images"
        wsi["A1"].font = TITLE_FONT
        wsi["A1"].alignment = CENTER
        wsi.row_dimensions[1].height = 28
        for ci, h in enumerate(["Filename", "URL", "Page", "Size (px)", "Caption"], 1):
            wsi.cell(2, ci, h)
        hdr_row(wsi, 2, 5)
        for ri, img in enumerate(images, 3):
            wsi.cell(ri, 1, img.get("filename", ""))
            wsi.cell(ri, 2, img.get("url", ""))
            wsi.cell(ri, 3, img.get("page", ""))
            wsi.cell(ri, 4, f"{img.get('width','')} × {img.get('height','')}")
            wsi.cell(ri, 5, img.get("caption", ""))
            data_row(wsi, ri, 5, alt=(ri % 2 == 0))
            wsi.row_dimensions[ri].height = 18
        auto_col(wsi)

    # ── Full Text sheet ──────────────────────────────────
    wst = wb.create_sheet("Full Text")
    wst.sheet_view.showGridLines = False
    wst.merge_cells("A1:B1")
    wst["A1"] = "Extracted Full Text"
    wst["A1"].font = TITLE_FONT
    wst["A1"].alignment = CENTER
    wst.row_dimensions[1].height = 28

    for ri, para in enumerate([p.strip() for p in (d.get("full_text") or "").split("\n") if p.strip()], 2):
        wst.merge_cells(f"A{ri}:B{ri}")
        cell = wst[f"A{ri}"]
        cell.value = para
        cell.alignment = WRAP
        cell.border = BORDER
        wst.row_dimensions[ri].height = 18
        if para.startswith("---") and para.endswith("---"):
            cell.font = Font(bold=True, color="1D4ED8")
            cell.fill = ALT_FILL

    wst.column_dimensions["A"].width = 110
    wst.column_dimensions["B"].width = 5

    # ── Save ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = filename.rsplit(".", 1)[0] + "_extracted.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get("/{job_id}/export/txt")
async def export_txt(job_id: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT filename, full_text FROM results WHERE job_id=?", (job_id,)
        ) as cur:
            row = await cur.fetchone()
    if not row:
        raise HTTPException(404, "Result not found")
    fname = (row["filename"] or job_id).rsplit(".", 1)[0] + "_text.txt"
    return StreamingResponse(
        io.StringIO(row["full_text"] or ""),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )
